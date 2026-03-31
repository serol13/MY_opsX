import streamlit as st
import pandas as pd
import io
import base64
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date, timezone, timedelta
TZ_GMT8 = timezone(timedelta(hours=8))
def now8(): return datetime.now(TZ_GMT8)
def fmt_ts(ts: str) -> str:
    """Convert ISO timestamp to dd/mm/yy hh:mm AM/PM"""
    if not ts: return "—"
    try:
        for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(str(ts)[:25], fmt)
                return dt.strftime("%d/%m/%y %I:%M %p")
            except:
                continue
        return str(ts)[:16].replace("T"," ")
    except:
        return str(ts)[:16]
import uuid
import plotly.express as px

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Operation Excellence Tracker", layout="wide",
                   initial_sidebar_state="expanded")

# ─────────────────────────────────────────────────────────────────────────────
# GITHUB CONFIG
# ─────────────────────────────────────────────────────────────────────────────
GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]
GITHUB_REPO  = st.secrets["GITHUB_REPO"]
FILE_PATH    = "tickets.csv"
REC_PATH     = "recurring.csv"
ACT_PATH     = "activity_log.csv"
BRANCH       = "main"
GITHUB_API   = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{FILE_PATH}"
REC_API      = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{REC_PATH}"
ACT_API      = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{ACT_PATH}"
ACT_COLS     = ["timestamp", "date", "username", "category", "description", "duration_min"]
GH_HEADERS   = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github.v3+json",
}

# ─────────────────────────────────────────────────────────────────────────────
# USERS
# ─────────────────────────────────────────────────────────────────────────────
USERS: dict = dict(st.secrets.get("users", {}))

# ─────────────────────────────────────────────────────────────────────────────
# CSV SCHEMA
# ─────────────────────────────────────────────────────────────────────────────
CSV_COLS = [
    "timestamp", "action", "ticket_id", "title", "platform", "priority",
    "status", "progress", "requestor", "due_date", "tags", "description",
    "updated_by", "notes", "complexity", "assigned_to", "image",
]

# ─────────────────────────────────────────────────────────────────────────────
# GITHUB HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def gh_load() -> tuple[pd.DataFrame, str | None]:
    r = requests.get(GITHUB_API, headers=GH_HEADERS)
    if r.status_code == 404:
        return pd.DataFrame(columns=CSV_COLS), None
    r.raise_for_status()
    data    = r.json()
    content = base64.b64decode(data["content"]).decode("utf-8")
    df = pd.read_csv(io.StringIO(content), dtype=str).fillna("")
    for col in CSV_COLS:
        if col not in df.columns:
            df[col] = ""
    return df[CSV_COLS], data["sha"]

def gh_append(new_row: dict) -> None:
    df, sha = gh_load()
    updated = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    csv_bytes = updated.to_csv(index=False).encode("utf-8")
    payload = {
        "message": (f"[{new_row['action']}] {new_row['ticket_id']} "
                    f"by {new_row['updated_by']} — {new_row['timestamp'][:16]}"),
        "content": base64.b64encode(csv_bytes).decode("utf-8"),
        "branch":  BRANCH,
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(GITHUB_API, headers=GH_HEADERS, json=payload)
    r.raise_for_status()
    st.session_state.log_df = updated

def current_tickets(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=CSV_COLS)
    active  = df[df["action"].isin(["CREATED", "UPDATED"])].copy()
    if active.empty:
        return pd.DataFrame(columns=CSV_COLS)
    active["progress"] = pd.to_numeric(active["progress"], errors="coerce").fillna(0).astype(int)
    latest  = (active.sort_values("timestamp")
                     .groupby("ticket_id", as_index=False)
                     .last())
    deleted = set(df[df["action"] == "DELETED"]["ticket_id"].tolist())
    return latest[~latest["ticket_id"].isin(deleted)].reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# RECURRING TASKS
# ─────────────────────────────────────────────────────────────────────────────
REC_COLS = ["task_id","title","description","frequency","day_info",
            "assigned_to","platform","created_by","created_at","active"]

def rec_load() -> tuple:
    r = requests.get(REC_API, headers=GH_HEADERS)
    if r.status_code == 404:
        return pd.DataFrame(columns=REC_COLS), None
    r.raise_for_status()
    data    = r.json()
    content = base64.b64decode(data["content"]).decode("utf-8")
    df = pd.read_csv(io.StringIO(content), dtype=str).fillna("")
    for col in REC_COLS:
        if col not in df.columns:
            df[col] = ""
    return df[REC_COLS], data["sha"]

def rec_save(df: pd.DataFrame, sha, msg: str) -> None:
    csv_bytes = df.to_csv(index=False).encode("utf-8")
    payload = {"message": msg,
               "content": base64.b64encode(csv_bytes).decode("utf-8"),
               "branch": BRANCH}
    if sha:
        payload["sha"] = sha
    r = requests.put(REC_API, headers=GH_HEADERS, json=payload)
    r.raise_for_status()
    st.session_state.rec_df = df

def act_load() -> tuple:
    r = requests.get(ACT_API, headers=GH_HEADERS)
    if r.status_code == 404:
        return pd.DataFrame(columns=ACT_COLS), None
    r.raise_for_status()
    data    = r.json()
    content = base64.b64decode(data["content"]).decode("utf-8")
    df = pd.read_csv(io.StringIO(content), dtype=str).fillna("")
    for col in ACT_COLS:
        if col not in df.columns:
            df[col] = ""
    return df[ACT_COLS], data["sha"]

def act_append(new_row: dict) -> None:
    df, sha = act_load()
    updated = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    csv_bytes = updated.to_csv(index=False).encode("utf-8")
    payload = {
        "message": f"[ACTIVITY] {new_row['username']} — {new_row['date']}",
        "content": base64.b64encode(csv_bytes).decode("utf-8"),
        "branch":  BRANCH,
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(ACT_API, headers=GH_HEADERS, json=payload)
    r.raise_for_status()
    st.session_state.act_df = updated

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE BOOTSTRAP
# ─────────────────────────────────────────────────────────────────────────────
if "log_df" not in st.session_state:
    with st.spinner("Loading from GitHub..."):
        try:
            st.session_state.log_df, _ = gh_load()
        except Exception as e:
            st.error(f"Could not reach GitHub: {e}")
            st.session_state.log_df = pd.DataFrame(columns=CSV_COLS)

if "logged_in_user" not in st.session_state:
    st.session_state.logged_in_user = None

if "show_login_form" not in st.session_state:
    st.session_state.show_login_form = False

if "app_unlocked" not in st.session_state:
    st.session_state.app_unlocked = False

if "rec_df" not in st.session_state:
    try:
        st.session_state.rec_df, _ = rec_load()
    except:
        st.session_state.rec_df = pd.DataFrame(columns=REC_COLS)

if "my_tasks_mode" not in st.session_state:
    st.session_state.my_tasks_mode = False

if "act_df" not in st.session_state:
    try:
        st.session_state.act_df, _ = act_load()
    except:
        st.session_state.act_df = pd.DataFrame(columns=ACT_COLS)

if "nav_page" not in st.session_state:
    st.session_state.nav_page = "Dashboard"

# NEW: jump-to-ticket state
if "jump_to_ticket" not in st.session_state:
    st.session_state.jump_to_ticket = None

# ── Unified Login Wall ────────────────────────────────────────────────────────
GUEST_PIN = str(st.secrets.get("GUEST_PIN", ""))

if "login_mode" not in st.session_state:
    st.session_state.login_mode = None

if GUEST_PIN and not st.session_state.app_unlocked:
    st.markdown("""
    <style>
    .login-wrap{max-width:420px;margin:60px auto 0}
    .login-card{background:#fff;border-radius:12px;border-top:6px solid #FFCC00;
                padding:36px 32px 28px;box-shadow:0 6px 28px rgba(0,0,0,.12)}
    .login-brand{font-size:30px;font-weight:900;color:#D40511;letter-spacing:1px;margin-bottom:2px}
    .login-title{font-size:20px;font-weight:700;color:#1A1A1A;margin-bottom:4px}
    .login-sub{font-size:13px;color:#6B6B6B;margin-bottom:28px}
    </style>
    <div class="login-wrap">
      <div class="login-card">
        <div class="login-brand">DHL</div>
        <div class="login-title">Operation Excellence Tracker</div>
        <div class="login-sub">Select how you want to access the app</div>
      </div>
    </div>""", unsafe_allow_html=True)

    _, center, _ = st.columns([1, 2, 1])
    with center:
        st.markdown("")

        if st.session_state.login_mode is None:
            col_a, col_g = st.columns(2)
            with col_a:
                if st.button("🔐  Admin Login", use_container_width=True):
                    st.session_state.login_mode = "admin"
                    st.rerun()
            with col_g:
                if st.button("👤  Guest Access", use_container_width=True):
                    st.session_state.login_mode = "guest"
                    st.rerun()

        elif st.session_state.login_mode == "admin":
            st.markdown("#### 🔐 Admin Login")
            with st.form("entry_admin_form"):
                uname = st.selectbox("Username", list(USERS.keys()))
                pin   = st.text_input("PIN", type="password", placeholder="Enter your PIN")
                c1, c2 = st.columns(2)
                with c1: ok   = st.form_submit_button("Sign In", use_container_width=True)
                with c2: back = st.form_submit_button("← Back",  use_container_width=True)
            if back:
                st.session_state.login_mode = None
                st.rerun()
            if ok:
                if uname in USERS and USERS[uname] == pin:
                    st.session_state.logged_in_user = uname
                    st.session_state.app_unlocked   = True
                    st.session_state.login_mode      = None
                    st.rerun()
                else:
                    st.error("Incorrect PIN. Please try again.")

        elif st.session_state.login_mode == "guest":
            st.markdown("#### 👤 Guest Access")
            st.caption("Enter the shared guest PIN to view the tracker.")
            with st.form("entry_guest_form"):
                entered = st.text_input("Guest PIN", type="password", placeholder="Enter guest PIN")
                c1, c2 = st.columns(2)
                with c1: ok   = st.form_submit_button("Unlock App", use_container_width=True)
                with c2: back = st.form_submit_button("← Back",     use_container_width=True)
            if back:
                st.session_state.login_mode = None
                st.rerun()
            if ok:
                if entered == GUEST_PIN:
                    st.session_state.app_unlocked = True
                    st.session_state.login_mode   = None
                    st.rerun()
                else:
                    st.error("Incorrect PIN. Please try again.")

    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# COLOURS & CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DHL_YELLOW = "#FFCC00"
DHL_RED    = "#D40511"
DHL_DARK   = "#1A1A1A"
DHL_GRAY   = "#6B6B6B"
DHL_LIGHT  = "#F5F5F5"
DHL_BORDER = "#E0E0E0"
DHL_WHITE  = "#FFFFFF"

PLATFORM_COLORS = {"Splunk": DHL_RED, "Power BI": "#0078D4", "Others": DHL_GRAY}
STATUS_COLORS   = {
    "Backlog":     "#9E9E9E",
    "In Progress": "#0078D4",
    "In Review":   "#FF8C00",
    "Done":        "#2E7D32",
    "Blocked":     DHL_RED,
}
PRIORITY_COLORS = {
    "Low":      "#2E7D32",
    "Medium":   "#FF8C00",
    "High":     DHL_RED,
    "Critical": "#6A0DAD",
}
PRIORITY_ORDER   = ["Low", "Medium", "High", "Critical"]
COMPLEXITY_ORDER = ["Simple", "Medium", "Complex", "Critical"]
COMPLEXITY_COLORS = {
    "Simple":   "#2E7D32",
    "Medium":   "#0078D4",
    "Complex":  "#FF8C00",
    "Critical": "#6A0DAD",
}
STATUS_ORDER   = ["Backlog", "In Progress", "In Review", "Blocked", "Done"]

def badge(label, bg, fg="#fff"):
    return (f'<span style="background:{bg};color:{fg};padding:3px 12px;'
            f'border-radius:4px;font-size:12px;font-weight:700">{label}</span>')

def progress_bar(pct, color=DHL_YELLOW):
    pct = int(float(pct)) if str(pct).replace('.','',1).isdigit() else 0
    return (
        '<div style="background:#E0E0E0;border-radius:4px;height:12px;'
        'width:100%;overflow:hidden;margin:6px 0 2px">'
        '<div style="background:' + color + ';width:' + str(pct) + '%;height:100%;border-radius:4px"></div></div>'
        '<small style="color:#6B6B6B;font-size:12px">' + str(pct) + '% complete</small>'
    )

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL NOTIFICATION
# ─────────────────────────────────────────────────────────────────────────────
def send_new_ticket_email(ticket: dict) -> None:
    try:
        gmail_user     = st.secrets["GMAIL_USER"]
        gmail_password = st.secrets["GMAIL_APP_PASSWORD"]
        notify_email   = st.secrets["NOTIFY_EMAIL"]
    except KeyError:
        return

    subject = f"[Ops Excellence Tracker] New Ticket: {ticket['ticket_id']} — {ticket['title']}"

    html = f"""
    <html><body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:20px">
    <div style="max-width:600px;margin:auto;background:#fff;border-radius:8px;
                border-top:5px solid #FFCC00;padding:28px 32px;box-shadow:0 2px 8px rgba(0,0,0,.08)">
      <div style="font-size:22px;font-weight:900;color:#D40511;letter-spacing:1px;margin-bottom:4px">DHL</div>
      <div style="font-size:18px;font-weight:700;color:#1A1A1A;margin-bottom:20px">New QA Ticket Submitted</div>
      <table style="width:100%;border-collapse:collapse;font-size:14px">
        <tr style="background:#FFCC00">
          <td style="padding:10px 14px;font-weight:700;color:#1A1A1A;width:35%">Ticket ID</td>
          <td style="padding:10px 14px;font-weight:700;color:#1A1A1A">{ticket['ticket_id']}</td>
        </tr>
        <tr style="background:#f9f9f9">
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Title</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['title']}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Platform</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['platform']}</td>
        </tr>
        <tr style="background:#f9f9f9">
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Priority</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['priority']}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Requestor</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['requestor']}</td>
        </tr>
        <tr style="background:#f9f9f9">
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Due Date</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['due_date']}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Submitted By</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['updated_by']}</td>
        </tr>
        <tr style="background:#f9f9f9">
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Description</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['description']}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Notes</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket.get('notes','—')}</td>
        </tr>
        <tr style="background:#f9f9f9">
          <td style="padding:10px 14px;color:#6B6B6B;font-weight:600">Submitted At</td>
          <td style="padding:10px 14px;color:#1A1A1A">{ticket['timestamp'][:16].replace('T', ' at ')}</td>
        </tr>
      </table>
      <div style="margin-top:24px;font-size:12px;color:#9E9E9E;border-top:1px solid #E0E0E0;padding-top:14px">
        This is an automated notification from your Operation Excellence Tracker.
      </div>
    </div>
    </body></html>
    """

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = gmail_user
    msg["To"]      = notify_email
    msg.attach(MIMEText(html, "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, notify_email, msg.as_string())

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(log_df: pd.DataFrame) -> bytes:
    def h(c): return c.lstrip("#")
    def thin():
        s = Side(style="thin", color="E0E0E0")
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr(ws, row, col, val, bg, fg="FFFFFF", merge_to=None, ht=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(bold=True, color=fg, name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        if merge_to:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(merge_to)}{row}")
        if ht:
            ws.row_dimensions[row].height = ht
        return c

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Audit Log"
    ws1.sheet_view.showGridLines = False

    hdr(ws1, 1, 1, "OPERATION EXCELLENCE TRACKER — FULL AUDIT LOG",
        h(DHL_YELLOW), h(DHL_DARK), merge_to=len(CSV_COLS), ht=34)
    ws1["A1"].font = Font(bold=True, color=h(DHL_DARK), name="Arial", size=13)
    hdr(ws1, 2, 1,
        f"Generated: {now8().strftime('%d %B %Y %H:%M')}  |  Total rows: {len(log_df)}",
        h(DHL_DARK), h(DHL_YELLOW), merge_to=len(CSV_COLS), ht=18)
    ws1["A2"].font = Font(italic=True, color=h(DHL_YELLOW), name="Arial", size=10)
    ws1.row_dimensions[3].height = 5

    col_widths = [20, 10, 14, 32, 12, 12, 14, 10, 16, 14, 20, 40, 16, 30]
    for i, (col, w) in enumerate(zip(CSV_COLS, col_widths), 1):
        hdr(ws1, 4, i, col.upper().replace("_", " "), h(DHL_RED), "FFFFFF")
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[4].height = 22
    ws1.freeze_panes = "A5"

    action_colors = {"CREATED": "2E7D32", "UPDATED": "0078D4", "DELETED": "D40511"}
    for r, (_, row) in enumerate(log_df.iterrows(), 5):
        bg = "FFFFFF" if r % 2 == 1 else "F5F5F5"
        ws1.row_dimensions[r].height = 18
        for ci, col in enumerate(CSV_COLS, 1):
            val = row.get(col, "")
            cell = ws1.cell(row=r, column=ci, value=val)
            cell.border = thin()
            cell.font   = Font(name="Arial", size=10, color=h(DHL_DARK))
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 12))
            if ci == 2:
                ac = action_colors.get(str(val), h(DHL_GRAY))
                cell.fill = PatternFill("solid", start_color=ac)
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2 = wb.create_sheet("Current Tickets")
    ws2.sheet_view.showGridLines = False

    latest = current_tickets(log_df)
    hdr(ws2, 1, 1, "CURRENT TICKET STATUS",
        h(DHL_YELLOW), h(DHL_DARK), merge_to=11, ht=34)
    ws2["A1"].font = Font(bold=True, color=h(DHL_DARK), name="Arial", size=13)
    hdr(ws2, 2, 1,
        f"Snapshot: {now8().strftime('%d %B %Y %H:%M')}  |  Active tickets: {len(latest)}",
        h(DHL_DARK), h(DHL_YELLOW), merge_to=11, ht=18)
    ws2["A2"].font = Font(italic=True, color=h(DHL_YELLOW), name="Arial", size=10)
    ws2.row_dimensions[3].height = 5

    show_cols   = ["ticket_id","title","platform","priority","status",
                   "progress","requestor","due_date","tags","updated_by","timestamp"]
    show_widths = [14, 34, 12, 12, 14, 11, 16, 14, 22, 16, 20]
    for i, (col, w) in enumerate(zip(show_cols, show_widths), 1):
        hdr(ws2, 4, i, col.upper().replace("_"," "), h(DHL_RED), "FFFFFF")
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[4].height = 22
    ws2.freeze_panes = "A5"

    for r, (_, row) in enumerate(latest.iterrows(), 5):
        bg = "FFFFFF" if r % 2 == 1 else "F5F5F5"
        ws2.row_dimensions[r].height = 18
        for ci, col in enumerate(show_cols, 1):
            val = row.get(col, "")
            cell = ws2.cell(row=r, column=ci, value=str(val))
            cell.border = thin()
            cell.font   = Font(name="Arial", size=10, color=h(DHL_DARK))
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center")
            if col == "platform":
                cell.fill = PatternFill("solid", start_color=h(PLATFORM_COLORS.get(val, DHL_GRAY)))
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "priority":
                cell.fill = PatternFill("solid", start_color=h(PRIORITY_COLORS.get(val, DHL_GRAY)))
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "status":
                cell.fill = PatternFill("solid", start_color=h(STATUS_COLORS.get(val, DHL_GRAY)))
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "progress":
                cell.value = int(val) if str(val).isdigit() else 0
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", start_color=bg)

    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [22, 12, 14, 14]):
        ws3.column_dimensions[col].width = w

    hdr(ws3, 1, 1, "SUMMARY", h(DHL_YELLOW), h(DHL_DARK), merge_to=4, ht=30)
    ws3["A1"].font = Font(bold=True, color=h(DHL_DARK), name="Arial", size=13)
    hdr(ws3, 2, 1, f"As of {now8().strftime('%d %B %Y')}",
        h(DHL_DARK), h(DHL_YELLOW), merge_to=4, ht=18)
    ws3["A2"].font = Font(italic=True, color=h(DHL_YELLOW), name="Arial", size=10)

    def summary_block(ws, start, title, counts, color_map, order):
        hdr(ws, start, 1, title, h(DHL_RED), "FFFFFF", merge_to=4, ht=20)
        for ci, lbl in enumerate(["Category", "Count", "% of Total"], 1):
            hdr(ws, start+1, ci, lbl, h(DHL_DARK), "FFFFFF")
        ws.row_dimensions[start+1].height = 18
        r = start + 2
        total = sum(counts.values()) or 1
        for key in order:
            cnt = counts.get(key, 0)
            ca = ws.cell(row=r, column=1, value=key)
            ca.fill = PatternFill("solid", start_color=h(color_map.get(key, DHL_GRAY)))
            ca.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            ca.alignment = Alignment(horizontal="center", vertical="center")
            ca.border = thin()
            cb = ws.cell(row=r, column=2, value=cnt)
            cb.font = Font(name="Arial", size=10, bold=True, color=h(DHL_DARK))
            cb.alignment = Alignment(horizontal="center"); cb.border = thin()
            cb.fill = PatternFill("solid", start_color="FFFFFF")
            cc = ws.cell(row=r, column=3, value=round(cnt/total*100, 1))
            cc.number_format = '0.0"%"'
            cc.font = Font(name="Arial", size=10, color=h(DHL_DARK))
            cc.alignment = Alignment(horizontal="center"); cc.border = thin()
            cc.fill = PatternFill("solid", start_color="F5F5F5")
            ws.row_dimensions[r].height = 20
            r += 1
        ct = ws.cell(row=r, column=1, value="TOTAL")
        ct.fill = PatternFill("solid", start_color=h(DHL_DARK))
        ct.font = Font(bold=True, color="FFFFFF", name="Arial")
        ct.alignment = Alignment(horizontal="center"); ct.border = thin()
        ctv = ws.cell(row=r, column=2, value=sum(counts.values()))
        ctv.fill = PatternFill("solid", start_color=h(DHL_YELLOW))
        ctv.font = Font(bold=True, name="Arial", color=h(DHL_DARK))
        ctv.alignment = Alignment(horizontal="center"); ctv.border = thin()
        ws.row_dimensions[r].height = 22
        return r + 2

    if not latest.empty:
        sc = dict(latest["status"].value_counts())
        pc = dict(latest["priority"].value_counts())
        pl = dict(latest["platform"].value_counts())
        nr = summary_block(ws3, 4,  "BY STATUS",   sc, STATUS_COLORS,   STATUS_ORDER)
        nr = summary_block(ws3, nr, "BY PRIORITY", pc, PRIORITY_COLORS, PRIORITY_ORDER)
        summary_block(ws3, nr, "BY PLATFORM", pl, PLATFORM_COLORS, ["Splunk","Power BI","Others"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;700&display=swap');
html,body,[class*="css"],.stApp{{font-family:'Roboto',sans-serif!important;background:{DHL_LIGHT}!important;color:{DHL_DARK}!important}}
.main .block-container{{background:{DHL_LIGHT};padding-top:1.2rem!important;max-width:1400px}}
[data-testid="stSidebar"]{{background:{DHL_DARK}!important;border-right:4px solid {DHL_YELLOW}!important}}
[data-testid="stSidebar"] p,[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,[data-testid="stSidebar"] div,[data-testid="stSidebar"] a{{color:{DHL_WHITE}!important}}
[data-testid="stSidebar"] hr{{border-color:#444!important}}
.dhl-topbar{{background:{DHL_YELLOW};padding:12px 20px;border-radius:8px;margin-bottom:20px;border-left:6px solid {DHL_RED}}}
.dhl-topbar h1{{margin:0;font-size:20px;font-weight:700;color:{DHL_DARK};text-transform:uppercase;letter-spacing:.5px}}
.dhl-topbar span{{font-size:12px;color:#555}}
.metric-card{{background:{DHL_WHITE};border:1px solid {DHL_BORDER};border-top:4px solid {DHL_YELLOW};border-radius:8px;padding:18px 20px;margin-bottom:8px}}
.metric-card .val{{font-size:2.2rem;font-weight:700;color:{DHL_DARK};line-height:1;margin-bottom:4px}}
.metric-card .lbl{{font-size:12px;font-weight:600;color:{DHL_GRAY};text-transform:uppercase;letter-spacing:.06em}}
.metric-card.red{{border-top-color:{DHL_RED}}}
.metric-card.blue{{border-top-color:#0078D4}}
.metric-card.green{{border-top-color:#2E7D32}}
.metric-card.orange{{border-top-color:#FF8C00}}
.section-header{{font-size:17px;font-weight:700;color:{DHL_DARK};border-left:5px solid {DHL_YELLOW};padding-left:12px;margin:24px 0 14px;text-transform:uppercase}}
.ticket-card{{background:{DHL_WHITE};border:1px solid {DHL_BORDER};border-left:5px solid {DHL_YELLOW};border-radius:8px;padding:15px 18px;margin-bottom:10px}}
.ticket-card:hover{{border-left-color:{DHL_RED}}}
.ticket-id{{font-size:11px;font-weight:700;color:{DHL_GRAY};text-transform:uppercase;letter-spacing:.08em}}
.ticket-title{{font-size:15px;font-weight:700;color:{DHL_DARK};margin:4px 0 8px}}
.pills{{display:flex;gap:8px;flex-wrap:wrap;margin:6px 0 8px}}
.readonly-banner{{background:#FFF9E6;border:1px solid {DHL_YELLOW};border-radius:6px;padding:8px 14px;font-size:13px;color:#7a6000;margin-bottom:12px}}
.stButton>button{{background:{DHL_YELLOW}!important;color:{DHL_DARK}!important;border:none!important;border-radius:5px!important;font-weight:700!important;font-size:14px!important;padding:9px 22px!important}}
.stButton>button:hover{{background:{DHL_RED}!important;color:{DHL_WHITE}!important}}
.stDownloadButton>button{{background:{DHL_YELLOW}!important;color:{DHL_DARK}!important;border:2px solid {DHL_YELLOW}!important;border-radius:5px!important;font-weight:700!important;font-size:13px!important;padding:9px 22px!important;width:100%!important}}
.stDownloadButton>button:hover{{background:{DHL_RED}!important;color:{DHL_WHITE}!important;border-color:{DHL_RED}!important}}
.stTextInput>div>div>input,.stTextArea>div>div>textarea{{background:{DHL_WHITE}!important;color:{DHL_DARK}!important;border:1px solid {DHL_BORDER}!important;border-radius:5px!important}}
.stSelectbox>div>div,.stMultiSelect>div>div{{background:{DHL_WHITE}!important;color:{DHL_DARK}!important}}
.stSlider>div>div>div>div{{background:{DHL_YELLOW}!important}}
.styled-ticket-table{{width:100%;border-collapse:collapse;font-size:13px}}
.styled-ticket-table th{{background:{DHL_YELLOW};color:{DHL_DARK};font-weight:700;padding:9px 12px;text-align:left;border:1px solid {DHL_BORDER};white-space:nowrap}}
.styled-ticket-table td{{padding:8px 12px;border:1px solid {DHL_BORDER};vertical-align:middle}}
.styled-ticket-table tr:nth-child(odd) td{{background:#FFFFFF}}
.styled-ticket-table tr:nth-child(even) td{{background:#F5F5F5}}
.styled-ticket-table tr:hover td{{background:#FFF9E6!important}}
.ticket-link-btn{{color:{DHL_RED};font-weight:700;font-size:12px;white-space:nowrap;font-family:'Roboto',sans-serif}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
user = st.session_state.logged_in_user

with st.sidebar:
    st.markdown("""
    <div style="padding:14px 0 6px">
      <div style="font-size:26px;font-weight:900;color:#FFCC00;letter-spacing:1px">DHL</div>
      <div style="font-size:12px;color:#aaa;margin-top:2px">Operation Excellence Tracker</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")
    if user:
        st.markdown(f'<div style="background:#FFCC00;color:#1A1A1A;padding:5px 14px;'
                    f'border-radius:20px;font-size:13px;font-weight:700">👤 {user}</div>',
                    unsafe_allow_html=True)
        st.markdown("")
        if st.button("Logout", use_container_width=True):
            st.session_state.logged_in_user  = None
            st.session_state.show_login_form = False
            st.session_state.app_unlocked    = False
            st.session_state.login_mode      = None
            st.rerun()
    else:
        st.markdown('<div style="font-size:12px;color:#aaa;margin-bottom:6px">Browsing as Guest</div>',
                    unsafe_allow_html=True)
        if not st.session_state.get("show_login_form"):
            if st.button("🔐 Admin Login", use_container_width=True):
                st.session_state.show_login_form = True
                st.rerun()
        else:
            with st.form("sidebar_admin_form"):
                uname = st.selectbox("Username", list(USERS.keys()))
                pin   = st.text_input("PIN", type="password", placeholder="Enter PIN")
                c1, c2 = st.columns(2)
                with c1: ok   = st.form_submit_button("Sign In")
                with c2: back = st.form_submit_button("Cancel")
            if back:
                st.session_state.show_login_form = False
                st.rerun()
            if ok:
                if uname in USERS and USERS[uname] == pin:
                    st.session_state.logged_in_user  = uname
                    st.session_state.show_login_form = False
                    st.rerun()
                else:
                    st.error("Incorrect PIN.")

    st.markdown("---")

    nav_options = ["Dashboard", "All Tickets", "Submit Request"]
    if user:
        nav_options += ["Recurring Tasks", "Update / Delete Ticket", "Activity Log"]

    if st.session_state.nav_page not in nav_options:
        st.session_state.nav_page = nav_options[0]

    selected = st.radio("Navigation", nav_options,
                        index=nav_options.index(st.session_state.nav_page),
                        label_visibility="collapsed")
    if selected != st.session_state.nav_page:
        st.session_state.nav_page      = selected
        st.session_state.my_tasks_mode = False
        st.rerun()

    st.markdown("---")

    log_df  = st.session_state.log_df
    tickets = current_tickets(log_df)
    total   = len(tickets)
    done    = len(tickets[tickets["status"] == "Done"]) if not tickets.empty else 0
    blocked = len(tickets[tickets["status"] == "Blocked"]) if not tickets.empty else 0

    if user and not tickets.empty:
        my_tasks = tickets[
            (tickets.get("assigned_to", pd.Series(dtype=str)) == user) &
            (~tickets["status"].isin(["Done"]))
        ] if "assigned_to" in tickets.columns else pd.DataFrame()
        my_count   = len(my_tasks)
        my_blocked = len(my_tasks[my_tasks["status"] == "Blocked"]) if not my_tasks.empty else 0
        blocked_line = f'<div style="font-size:11px;color:#D40511;margin-top:2px;font-weight:700">{my_blocked} blocked</div>' if my_blocked else ""
        st.markdown(f"""
        <div style="background:#1e1e1e;border:1px solid #FFCC00;border-radius:6px;
                    padding:10px 14px;margin-bottom:4px">
          <div style="font-size:11px;color:#FFCC00;font-weight:700;text-transform:uppercase;
                      letter-spacing:.06em;margin-bottom:4px">My Tasks</div>
          <div style="font-size:28px;font-weight:900;color:#FFCC00;line-height:1">{my_count}</div>
          <div style="font-size:11px;color:#aaa;margin-top:2px">pending tickets assigned to you</div>
          {blocked_line}
        </div>""", unsafe_allow_html=True)
        if my_count > 0:
            if st.button("View & Update My Tasks →", key="my_tasks_btn", use_container_width=True):
                st.session_state.my_tasks_mode = True
                st.session_state.nav_page = "Update / Delete Ticket"
                st.rerun()

    rec_df_sb = st.session_state.rec_df
    if user and not rec_df_sb.empty:
        my_rec = rec_df_sb[(rec_df_sb["assigned_to"] == user) & (rec_df_sb["active"] == "Yes")]
        if len(my_rec) > 0:
            daily_c   = len(my_rec[my_rec["frequency"] == "Daily"])
            weekly_c  = len(my_rec[my_rec["frequency"] == "Weekly"])
            monthly_c = len(my_rec[my_rec["frequency"] == "Monthly"])
            st.markdown(f"""
            <div style="background:#1e1e1e;border:1px solid #555;border-radius:6px;
                        padding:10px 14px;margin-bottom:8px">
              <div style="font-size:11px;color:#aaa;font-weight:700;text-transform:uppercase;
                          letter-spacing:.06em;margin-bottom:4px">My Recurring</div>
              <div style="font-size:12px;color:#ccc;line-height:2">
                {"<b style='color:#FFCC00'>D:</b> " + str(daily_c) + "&nbsp;&nbsp;" if daily_c else ""}{"<b style='color:#FFCC00'>W:</b> " + str(weekly_c) + "&nbsp;&nbsp;" if weekly_c else ""}{"<b style='color:#FFCC00'>M:</b> " + str(monthly_c) if monthly_c else ""}
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown(f"""
    <div style="font-size:13px;color:#ccc;line-height:2.2">
      <b style="color:#fff">Active tickets:</b> {total}<br>
      <b style="color:#FFCC00">Done:</b> {done}<br>
      <b style="color:#D40511">Blocked:</b> {blocked}
    </div>""", unsafe_allow_html=True)
    st.markdown("")

    if not log_df.empty:
        excel_data = build_excel(log_df)
        st.download_button(
            "Download Excel Report",
            data=excel_data,
            file_name=f"QA_Tracker_{now8().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    st.markdown("")
    if st.button("Refresh from GitHub"):
        with st.spinner("Syncing..."):
            try:
                st.session_state.log_df, _ = gh_load()
                st.rerun()
            except Exception as e:
                st.error(str(e))

# ─────────────────────────────────────────────────────────────────────────────
# PAGE ROUTING
# ─────────────────────────────────────────────────────────────────────────────
page    = st.session_state.nav_page
log_df  = st.session_state.log_df
tickets = current_tickets(log_df)

CHART = dict(
    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor=DHL_WHITE,
    font=dict(color=DHL_DARK, family="Roboto"),
    title_font=dict(color=DHL_DARK, size=14),
    margin=dict(t=40, b=20, l=10, r=10),
)

PAGE_META = {
    "Dashboard":              ("Dashboard Overview",        "Live summary of all Operation Excellence requests"),
    "All Tickets":            ("All Tickets",               "Browse and filter every submitted ticket"),
    "Recurring Tasks":        ("Recurring Tasks",           "Daily, weekly and monthly task reference"),
    "Submit Request":         ("Submit New Request",        "Anyone can raise a new request"),
    "Update / Delete Ticket": ("Update / Delete Ticket",   f"Editing as: {user}"),
    "Activity Log":           ("Activity Log",              f"Daily work log & contribution heatmap — {user}"),
}
title, subtitle = PAGE_META.get(page, ("Operation Excellence Tracker", ""))
st.markdown(f"""
<div class="dhl-topbar">
  <div><h1>{title}</h1><span>{subtitle}</span></div>
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
if page == "Dashboard":
    df = tickets

    c1,c2,c3,c4,c5 = st.columns(5)
    for col, label, val, cls in [
        (c1, "Total Tickets",  len(df),                                                  ""),
        (c2, "In Progress",    len(df[df.status=="In Progress"]) if not df.empty else 0, "blue"),
        (c3, "In Review",      len(df[df.status=="In Review"])   if not df.empty else 0, "orange"),
        (c4, "Blocked",        len(df[df.status=="Blocked"])     if not df.empty else 0, "red"),
        (c5, "Done",           len(df[df.status=="Done"])        if not df.empty else 0, "green"),
    ]:
        with col:
            st.markdown(f'<div class="metric-card {cls}"><div class="val">{val}</div>'
                        f'<div class="lbl">{label}</div></div>', unsafe_allow_html=True)

    if df.empty:
        st.info("No tickets yet. Submit the first request using the sidebar.")
    else:
        st.markdown("---")
        r1a, r1b = st.columns(2)
        with r1a:
            sc = df["status"].value_counts().reset_index()
            sc.columns = ["Status","Count"]
            fig = px.pie(sc, names="Status", values="Count", hole=0.55,
                         title="By Status", color="Status", color_discrete_map=STATUS_COLORS)
            fig.update_layout(**CHART)
            st.plotly_chart(fig, use_container_width=True)
        with r1b:
            pc = df["platform"].value_counts().reset_index()
            pc.columns = ["Platform","Count"]
            fig2 = px.bar(pc, x="Platform", y="Count", title="By Platform",
                          color="Platform", color_discrete_map=PLATFORM_COLORS, text="Count")
            fig2.update_layout(**CHART, showlegend=False,
                               xaxis=dict(gridcolor=DHL_BORDER),
                               yaxis=dict(gridcolor=DHL_BORDER))
            fig2.update_traces(textposition="outside", marker_line_width=0)
            st.plotly_chart(fig2, use_container_width=True)

        r2a, r2b = st.columns(2)
        with r2a:
            prc = df["priority"].value_counts().reindex(PRIORITY_ORDER, fill_value=0).reset_index()
            prc.columns = ["Priority","Count"]
            fig3 = px.bar(prc, x="Priority", y="Count", title="By Priority",
                          color="Priority", color_discrete_map=PRIORITY_COLORS, text="Count")
            fig3.update_layout(**CHART, showlegend=False,
                               xaxis=dict(gridcolor=DHL_BORDER),
                               yaxis=dict(gridcolor=DHL_BORDER))
            fig3.update_traces(textposition="outside", marker_line_width=0)
            st.plotly_chart(fig3, use_container_width=True)
        with r2b:
            df["progress"] = pd.to_numeric(df["progress"], errors="coerce").fillna(0)
            ap = df.groupby("status")["progress"].mean().reset_index()
            ap.columns = ["Status","Avg"]
            fig4 = px.bar(ap, x="Status", y="Avg", title="Avg Progress % by Status",
                          color="Status", color_discrete_map=STATUS_COLORS,
                          text=ap["Avg"].round(1).astype(str)+"%")
            fig4.update_layout(**CHART, showlegend=False,
                               xaxis=dict(gridcolor=DHL_BORDER),
                               yaxis=dict(gridcolor=DHL_BORDER, range=[0,115]))
            fig4.update_traces(textposition="outside", marker_line_width=0)
            st.plotly_chart(fig4, use_container_width=True)

        if not log_df.empty:
            st.markdown('<div class="section-header">Activity Timeline</div>', unsafe_allow_html=True)
            tl = log_df.copy()
            tl["date"] = tl["timestamp"].str[:10]
            daily = tl.groupby(["date","action"]).size().reset_index(name="count")
            fig5 = px.bar(daily, x="date", y="count", color="action",
                          title="Daily Actions",
                          color_discrete_map={"CREATED":"#2E7D32","UPDATED":"#0078D4","DELETED":"#D40511"})
            fig5.update_layout(**CHART, xaxis=dict(gridcolor=DHL_BORDER),
                               yaxis=dict(gridcolor=DHL_BORDER))
            st.plotly_chart(fig5, use_container_width=True)

        st.markdown('<div class="section-header">Recent Tickets</div>', unsafe_allow_html=True)
        recent = df.sort_values("timestamp", ascending=False).head(5)
        for _, t in recent.iterrows():
            pct = int(float(t.get("progress", 0) or 0))
            bc  = STATUS_COLORS.get(t["status"], DHL_YELLOW)
            tag_html = " ".join(
                f'<span style="background:{DHL_LIGHT};color:{DHL_DARK};border:1px solid {DHL_BORDER};'
                f'padding:1px 8px;border-radius:4px;font-size:11px;font-weight:600">{tg}</span>'
                for tg in str(t.get("tags","")).split(",") if tg.strip()
            )
            complexity_badge = badge("Complexity: " + str(t.get("complexity","")), "#555555") if t.get("complexity") else ""
            assigned_badge   = badge("Assigned: " + str(t.get("assigned_to","")), "#1A1A1A", DHL_YELLOW) if t.get("assigned_to") else ""
            st.markdown(
                '<div class="ticket-card">'
                + f'<div class="ticket-id">{t["ticket_id"]} · {t.get("requestor","")} · Due {t.get("due_date","—")}</div>'
                + f'<div class="ticket-title">{t["title"]}</div>'
                + '<div class="pills">'
                + badge(t["platform"], PLATFORM_COLORS.get(t["platform"], DHL_GRAY))
                + badge(t["status"],   STATUS_COLORS.get(t["status"],    DHL_GRAY))
                + badge(t["priority"], PRIORITY_COLORS.get(t["priority"], DHL_GRAY))
                + complexity_badge + assigned_badge + tag_html + '</div>'
                + progress_bar(pct, bc)
                + '</div>',
                unsafe_allow_html=True)

        rec_df_dash = st.session_state.rec_df
        active_rec  = rec_df_dash[rec_df_dash["active"] == "Yes"] if not rec_df_dash.empty else pd.DataFrame()
        if not active_rec.empty:
            st.markdown('<div class="section-header">Recurring Tasks</div>', unsafe_allow_html=True)
            FREQ_COLORS = {"Daily": "#2E7D32", "Weekly": "#0078D4", "Monthly": "#6A0DAD"}
            for freq in ["Daily", "Weekly", "Monthly"]:
                grp = active_rec[active_rec["frequency"] == freq]
                if grp.empty:
                    continue
                st.markdown(f'<div style="font-size:13px;font-weight:700;color:{FREQ_COLORS[freq]};'
                            f'margin:12px 0 6px;text-transform:uppercase;letter-spacing:.05em">'
                            f'{freq} ({len(grp)})</div>', unsafe_allow_html=True)
                for _, r in grp.iterrows():
                    assigned_b = badge(str(r.get("assigned_to","Unassigned")), "#1A1A1A", DHL_YELLOW) if r.get("assigned_to") else badge("Unassigned","#9E9E9E")
                    plat_b     = badge(str(r.get("platform","Others")), PLATFORM_COLORS.get(str(r.get("platform","")),"#9E9E9E"))
                    day_info   = f' · {r["day_info"]}' if r.get("day_info") else ""
                    st.markdown(
                        f'<div class="ticket-card" style="border-left-color:{FREQ_COLORS[freq]}">'
                        + f'<div class="ticket-id">{r["task_id"]}{day_info}</div>'
                        + f'<div class="ticket-title">{r["title"]}</div>'
                        + '<div class="pills">' + plat_b + badge(freq, FREQ_COLORS[freq]) + assigned_b + '</div>'
                        + (f'<p style="color:#6B6B6B;font-size:12px;margin-top:4px">{r["description"]}</p>' if r.get("description") else "")
                        + '</div>',
                        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: ALL TICKETS
# ─────────────────────────────────────────────────────────────────────────────
elif page == "All Tickets":
    if tickets.empty:
        st.info("No tickets yet.")
    else:
        fc1,fc2,fc3,fc4 = st.columns(4)
        with fc1: fp = st.multiselect("Platform", ["Splunk","Power BI","Others"],
                                      default=["Splunk","Power BI","Others"])
        with fc2: fs = st.multiselect("Status",   STATUS_ORDER,   default=STATUS_ORDER)
        with fc3: fr = st.multiselect("Priority", PRIORITY_ORDER, default=PRIORITY_ORDER)
        with fc4: fq = st.text_input("Search", placeholder="title / requestor...")

        df = tickets.copy()
        df = df[df["platform"].isin(fp) & df["status"].isin(fs) & df["priority"].isin(fr)]
        if fq:
            mask = (df["title"].str.contains(fq, case=False, na=False) |
                    df["requestor"].str.contains(fq, case=False, na=False))
            df = df[mask]

        sc1,sc2 = st.columns([2,1])
        with sc1:
            sort_by = st.selectbox("Sort by", ["Newest first","Oldest first",
                                               "Priority (high to low)",
                                               "Progress (high to low)","Due Date"])
        with sc2:
            view_mode = st.radio("View", ["Cards","Table"], horizontal=True)

        def skey(row):
            if sort_by == "Newest first":           return row["timestamp"]
            if sort_by == "Oldest first":           return row["timestamp"]
            if sort_by == "Priority (high to low)": return PRIORITY_ORDER.index(row["priority"]) if row["priority"] in PRIORITY_ORDER else 99
            if sort_by == "Progress (high to low)": return -int(row["progress"]) if str(row["progress"]).isdigit() else 0
            return row.get("due_date","")

        df["_sk"] = df.apply(skey, axis=1)
        df = df.sort_values("_sk", ascending=(sort_by not in ["Newest first","Progress (high to low)"]))
        st.caption(f"Showing {len(df)} ticket(s)")

        # ── CARDS VIEW ────────────────────────────────────────────────────────
        if view_mode == "Cards":
            for _, t in df.iterrows():
                pct = int(float(t.get("progress", 0) or 0))
                bc  = STATUS_COLORS.get(t["status"], DHL_YELLOW)
                tag_html = " ".join(
                    f'<span style="background:{DHL_LIGHT};color:{DHL_DARK};border:1px solid {DHL_BORDER};'
                    f'padding:1px 8px;border-radius:4px;font-size:11px;font-weight:600">{tg}</span>'
                    for tg in str(t.get("tags","")).split(",") if tg.strip()
                )
                desc       = str(t.get("description",""))
                desc_short = desc[:200] + ("..." if len(desc) > 200 else "")
                complexity_badge = badge("Complexity: " + str(t.get("complexity","")), "#555555") if t.get("complexity") else ""
                assigned_badge   = badge("Assigned: " + str(t.get("assigned_to","")), "#1A1A1A", DHL_YELLOW) if t.get("assigned_to") else ""
                st.markdown(
                    '<div class="ticket-card">'
                    + f'<div class="ticket-id">{t["ticket_id"]} · {t.get("requestor","")} · Due {t.get("due_date","—")} · Assigned: {t.get("assigned_to","Unassigned")} · Updated by {t.get("updated_by","")}</div>'
                    + f'<div class="ticket-title">{t["title"]}</div>'
                    + '<div class="pills">'
                    + badge(t["platform"], PLATFORM_COLORS.get(t["platform"], DHL_GRAY))
                    + badge(t["status"],   STATUS_COLORS.get(t["status"],    DHL_GRAY))
                    + badge(t["priority"], PRIORITY_COLORS.get(t["priority"], DHL_GRAY))
                    + complexity_badge + assigned_badge + tag_html + '</div>'
                    + progress_bar(pct, bc)
                    + f'<p style="color:#6B6B6B;font-size:13px;margin-top:6px">{desc_short}</p>'
                    + '</div>',
                    unsafe_allow_html=True)

        # ── TABLE VIEW ────────────────────────────────────────────────────────
        else:
            st.markdown("#### Column Filters")
            tf1, tf2, tf3, tf4, tf5 = st.columns(5)
            with tf1:
                all_assigned = sorted([x for x in df["assigned_to"].dropna().unique() if x])
                tf_assigned = st.multiselect("Assigned To", all_assigned, default=all_assigned, key="tf_assigned")
            with tf2:
                tf_complexity = st.multiselect("Complexity", COMPLEXITY_ORDER,
                                               default=[c for c in COMPLEXITY_ORDER if c in df["complexity"].values],
                                               key="tf_complexity")
            with tf3:
                tf_progress_min, tf_progress_max = st.select_slider(
                    "Progress %", options=list(range(0, 101, 5)),
                    value=(0, 100), key="tf_progress")
            with tf4:
                all_tags = sorted(set(
                    tag.strip() for tags in df["tags"].dropna()
                    for tag in str(tags).split(",") if tag.strip()
                ))
                tf_tags = st.multiselect("Tags", all_tags, key="tf_tags")
            with tf5:
                all_updaters = sorted([x for x in df["updated_by"].dropna().unique() if x])
                tf_updater = st.multiselect("Last Updated By", all_updaters, default=all_updaters, key="tf_updater")

            df_tbl = df.copy()
            df_tbl["progress"] = pd.to_numeric(df_tbl["progress"], errors="coerce").fillna(0).astype(int)
            if tf_assigned:
                df_tbl = df_tbl[df_tbl["assigned_to"].isin(tf_assigned) | (df_tbl["assigned_to"] == "")]
            if tf_complexity:
                df_tbl = df_tbl[df_tbl["complexity"].isin(tf_complexity) | (df_tbl["complexity"] == "")]
            df_tbl = df_tbl[
                (df_tbl["progress"] >= tf_progress_min) &
                (df_tbl["progress"] <= tf_progress_max)
            ]
            if tf_tags:
                def has_tag(tag_str):
                    row_tags = [t.strip() for t in str(tag_str).split(",")]
                    return any(t in row_tags for t in tf_tags)
                df_tbl = df_tbl[df_tbl["tags"].apply(has_tag)]
            if tf_updater:
                df_tbl = df_tbl[df_tbl["updated_by"].isin(tf_updater)]

            st.caption(f"Showing {len(df_tbl)} ticket(s)")

            # ── Helper badge functions for HTML table ─────────────────────────
            def status_cell(val):
                c = STATUS_COLORS.get(val, "#9E9E9E")
                return f'<span style="background:{c};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;white-space:nowrap">{val}</span>'

            def priority_cell(val):
                c = PRIORITY_COLORS.get(val, "#9E9E9E")
                return f'<span style="background:{c};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;white-space:nowrap">{val}</span>'

            def complexity_cell(val):
                if not val:
                    return '<span style="color:#9E9E9E">—</span>'
                c = COMPLEXITY_COLORS.get(val, "#9E9E9E")
                return f'<span style="background:{c};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;white-space:nowrap">{val}</span>'

            def platform_cell(val):
                c = PLATFORM_COLORS.get(val, DHL_GRAY)
                return f'<span style="background:{c};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;white-space:nowrap">{val}</span>'

            def progress_cell(val):
                pct = int(val) if str(val).replace('.','',1).isdigit() else 0
                bar = (f'<div style="background:#E0E0E0;border-radius:3px;height:8px;width:70px;'
                       f'display:inline-block;vertical-align:middle;margin-right:5px">'
                       f'<div style="background:{DHL_YELLOW};width:{pct}%;height:100%;border-radius:3px"></div></div>')
                return f'{bar}<span style="font-size:12px;color:{DHL_DARK}">{pct}%</span>'

            # ── Build HTML table ──────────────────────────────────────────────
            headers = ["Ticket ID", "Title", "Platform", "Priority", "Complexity",
                       "Status", "Progress", "Requestor", "Assigned To",
                       "Due Date", "Tags", "Last Updated By", "Last Updated At", "Latest Notes"]

            header_html = "".join(f"<th>{h}</th>" for h in headers)

            rows_html = ""
            for _, row in df_tbl.iterrows():
                tid        = str(row.get("ticket_id", ""))
                title_disp = str(row.get("title", ""))
                title_disp = title_disp[:55] + "…" if len(title_disp) > 55 else title_disp
                notes_disp = str(row.get("notes", ""))
                notes_disp = notes_disp[:70] + "…" if len(notes_disp) > 70 else notes_disp
                ts_disp    = fmt_ts(row.get("timestamp", ""))
                tags_disp  = str(row.get("tags", ""))

                rows_html += (
                    f"<tr>"
                    f'<td style="white-space:nowrap"><span class="ticket-link-btn">{tid}</span></td>'
                    f"<td>{title_disp}</td>"
                    f"<td>{platform_cell(row.get('platform',''))}</td>"
                    f"<td>{priority_cell(row.get('priority',''))}</td>"
                    f"<td>{complexity_cell(row.get('complexity',''))}</td>"
                    f"<td>{status_cell(row.get('status',''))}</td>"
                    f"<td style='min-width:130px'>{progress_cell(row.get('progress',0))}</td>"
                    f"<td style='white-space:nowrap'>{row.get('requestor','')}</td>"
                    f"<td style='white-space:nowrap'>{row.get('assigned_to','—')}</td>"
                    f"<td style='white-space:nowrap'>{row.get('due_date','—')}</td>"
                    f"<td style='font-size:11px;color:{DHL_GRAY}'>{tags_disp}</td>"
                    f"<td style='white-space:nowrap'>{row.get('updated_by','')}</td>"
                    f"<td style='white-space:nowrap'>{ts_disp}</td>"
                    f"<td style='font-size:12px;color:{DHL_GRAY}'>{notes_disp}</td>"
                    f"</tr>"
                )

            st.markdown(
                f'<div style="overflow-x:auto;border:1px solid {DHL_BORDER};border-radius:8px">'
                f'<table class="styled-ticket-table">'
                f'<thead><tr>{header_html}</tr></thead>'
                f'<tbody>{rows_html}</tbody>'
                f'</table></div>',
                unsafe_allow_html=True
            )

            # ── Jump buttons (admin only) — one per ticket ────────────────────
            if user and not df_tbl.empty:
                st.markdown("")
                st.markdown(
                    f'<div style="background:#FFF9E6;border:1px solid {DHL_YELLOW};border-radius:6px;'
                    f'padding:8px 14px;font-size:13px;color:#7a6000;margin-bottom:8px">'
                    f'🔗 Click a ticket button below to open it in the editor</div>',
                    unsafe_allow_html=True
                )
                # Render buttons in rows of 8
                tids = df_tbl["ticket_id"].tolist()
                cols_per_row = 8
                for row_start in range(0, len(tids), cols_per_row):
                    chunk = tids[row_start:row_start + cols_per_row]
                    btn_cols = st.columns(len(chunk))
                    for i, tid in enumerate(chunk):
                        with btn_cols[i]:
                            if st.button(tid, key=f"jump_{tid}"):
                                st.session_state.jump_to_ticket = tid
                                st.session_state.nav_page = "Update / Delete Ticket"
                                st.session_state.my_tasks_mode = False
                                st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: RECURRING TASKS
# ─────────────────────────────────────────────────────────────────────────────
elif page == "Recurring Tasks":
    rec_df   = st.session_state.rec_df
    FREQ_COLORS = {"Daily": "#2E7D32", "Weekly": "#0078D4", "Monthly": "#6A0DAD"}

    if user:
        with st.expander("+ Add New Recurring Task", expanded=False):
            rc1, rc2 = st.columns(2)
            with rc1:
                r_title    = st.text_input("Task Name *", key="r_title")
                r_freq     = st.selectbox("Frequency *", ["Daily","Weekly","Monthly"], key="r_freq")
                r_platform = st.selectbox("Platform", ["Splunk","Power BI","Others"], key="r_plat")
            with rc2:
                user_list  = list(USERS.keys())
                r_assigned = st.selectbox("Assign To *", user_list, key="r_assign")
                day_options = {
                    "Daily":   ["Every day"],
                    "Weekly":  ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
                    "Monthly": ["1st","2nd","3rd","4th","5th","10th","15th","20th","25th","Last day"],
                }
                r_day = st.selectbox("When", day_options[st.session_state.get("r_freq","Daily")], key="r_day")
            r_desc = st.text_area("Description (optional)", key="r_desc", height=80)
            if st.button("Add Recurring Task", key="add_rec"):
                if not r_title:
                    st.error("Task Name is required.")
                else:
                    new_id  = "REC-" + str(uuid.uuid4())[:6].upper()
                    new_row = {
                        "task_id":     new_id,
                        "title":       r_title,
                        "description": r_desc,
                        "frequency":   r_freq,
                        "day_info":    r_day,
                        "assigned_to": r_assigned,
                        "platform":    r_platform,
                        "created_by":  user,
                        "created_at":  now8().isoformat(timespec="seconds"),
                        "active":      "Yes",
                    }
                    with st.spinner("Saving..."):
                        try:
                            df_cur, sha = rec_load()
                            df_new = pd.concat([df_cur, pd.DataFrame([new_row])], ignore_index=True)
                            rec_save(df_new, sha, f"[ADD] {new_id} by {user}")
                            st.success(f"Recurring task {new_id} added!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Save failed: {e}")
    else:
        st.markdown('<div class="readonly-banner">Login to add or edit recurring tasks.</div>',
                    unsafe_allow_html=True)

    if rec_df.empty:
        st.info("No recurring tasks yet. Login and add the first one above.")
    else:
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            f_freq = st.multiselect("Frequency", ["Daily","Weekly","Monthly"],
                                    default=["Daily","Weekly","Monthly"], key="rec_f_freq")
        with fc2:
            all_assignees = sorted(rec_df["assigned_to"].dropna().unique().tolist())
            f_owner = st.multiselect("Assigned To", all_assignees,
                                     default=all_assignees, key="rec_f_owner")
        with fc3:
            f_active = st.radio("Show", ["Active only","All"], horizontal=True, key="rec_active")

        view_df = rec_df.copy()
        if f_freq:   view_df = view_df[view_df["frequency"].isin(f_freq)]
        if f_owner:  view_df = view_df[view_df["assigned_to"].isin(f_owner)]
        if f_active == "Active only":
            view_df = view_df[view_df["active"] == "Yes"]

        st.caption(f"Showing {len(view_df)} recurring task(s)")

        for freq in ["Daily","Weekly","Monthly"]:
            grp = view_df[view_df["frequency"] == freq]
            if grp.empty:
                continue
            st.markdown(f'<div class="section-header" style="border-left-color:{FREQ_COLORS[freq]}">'
                        f'{freq} Tasks ({len(grp)})</div>', unsafe_allow_html=True)
            for _, r in grp.iterrows():
                assigned_b = badge(str(r.get("assigned_to","Unassigned")), "#1A1A1A", DHL_YELLOW) if r.get("assigned_to") else badge("Unassigned","#9E9E9E")
                plat_b     = badge(str(r.get("platform","Others")), PLATFORM_COLORS.get(str(r.get("platform","")),"#9E9E9E"))
                active_b   = badge("Active","#2E7D32") if r.get("active") == "Yes" else badge("Inactive","#9E9E9E")
                day_info   = f' · {r["day_info"]}' if r.get("day_info") else ""

                col_card, col_btn = st.columns([5,1])
                with col_card:
                    st.markdown(
                        f'<div class="ticket-card" style="border-left-color:{FREQ_COLORS[freq]};margin-bottom:4px">'
                        + f'<div class="ticket-id">{r["task_id"]}{day_info} · Added by {r.get("created_by","")}</div>'
                        + f'<div class="ticket-title">{r["title"]}</div>'
                        + '<div class="pills">' + plat_b + badge(freq, FREQ_COLORS[freq]) + assigned_b + active_b + '</div>'
                        + (f'<p style="color:#6B6B6B;font-size:12px;margin-top:4px">{r["description"]}</p>' if r.get("description") else "")
                        + '</div>',
                        unsafe_allow_html=True)
                with col_btn:
                    if user:
                        st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
                        action = "Deactivate" if r.get("active") == "Yes" else "Reactivate"
                        if st.button(action, key=f"tog_{r['task_id']}"):
                            with st.spinner("Updating..."):
                                try:
                                    df_cur, sha = rec_load()
                                    df_cur.loc[df_cur["task_id"] == r["task_id"], "active"] = \
                                        "No" if r.get("active") == "Yes" else "Yes"
                                    rec_save(df_cur, sha, f"[TOGGLE] {r['task_id']} by {user}")
                                    st.rerun()
                                except Exception as e:
                                    st.error(str(e))
                        if st.button("Delete", key=f"del_{r['task_id']}"):
                            with st.spinner("Deleting..."):
                                try:
                                    df_cur, sha = rec_load()
                                    df_cur = df_cur[df_cur["task_id"] != r["task_id"]]
                                    rec_save(df_cur, sha, f"[DELETE] {r['task_id']} by {user}")
                                    st.rerun()
                                except Exception as e:
                                    st.error(str(e))

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: SUBMIT REQUEST
# ─────────────────────────────────────────────────────────────────────────────
elif page == "Submit Request":
    if not user:
        st.markdown(
            '<div class="readonly-banner">You are submitting as a guest. '
            'Login (top-left) to enable ticket updates and deletions.</div>',
            unsafe_allow_html=True)

    with st.form("submit_form", clear_on_submit=True):
        c1,c2 = st.columns(2)
        with c1:
            title_val    = st.text_input("Ticket Title *", placeholder="e.g. Sales Dashboard KPI refresh")
            platform_val = st.selectbox("Platform *", ["Splunk","Power BI","Others"])
            priority_val = st.selectbox("Priority *", PRIORITY_ORDER)
        with c2:
            default_name  = user if user else ""
            requestor_val = st.text_input("Your Name *", value=default_name,
                                          placeholder="Enter your name")
            requestor_email = st.text_input("Your Email (optional)", placeholder="e.g. you@example.com")
            due_val  = st.date_input("Target Due Date", value=date.today())
            tags_val = st.text_input("Tags (comma-separated)", placeholder="e.g. kpi, finance, Q2")
        desc_val  = st.text_area("Description / Requirements *",
                                 placeholder="Describe the request in detail...", height=150)
        notes_val = st.text_input("Notes (optional)", placeholder="Any additional notes for this submission...")
        submitted = st.form_submit_button("Submit Ticket")

    if submitted:
        missing = []
        if not title_val:     missing.append("Ticket Title")
        if not requestor_val: missing.append("Your Name")
        if not desc_val:      missing.append("Description / Requirements")
        if missing:
            st.error(f"Please fill in the following required fields: **{', '.join(missing)}**")
        else:
            tid        = "QA-" + str(uuid.uuid4())[:6].upper()
            updated_by = user if user else f"Guest:{requestor_val}"
            email_note = f"[Email: {requestor_email.strip()}] " if requestor_email.strip() else ""
            row = {
                "timestamp":   now8().isoformat(timespec="seconds"),
                "action":      "CREATED",
                "ticket_id":   tid,
                "title":       title_val,
                "platform":    platform_val,
                "priority":    priority_val,
                "status":      "Backlog",
                "progress":    0,
                "requestor":   requestor_val,
                "due_date":    str(due_val),
                "tags":        ", ".join([t.strip() for t in tags_val.split(",") if t.strip()]),
                "description": desc_val,
                "updated_by":  updated_by,
                "notes":       email_note + notes_val,
                "complexity":  "",
                "assigned_to": "",
            }
            with st.spinner("Saving to GitHub..."):
                try:
                    gh_append(row)
                    send_new_ticket_email(row)
                    st.success(f"Ticket **{tid}** submitted and logged to CSV.")
                    st.markdown(f"""<div class="ticket-card">
                      <div class="ticket-id">{tid}</div>
                      <div class="ticket-title">{title_val}</div>
                      <div class="pills">
                        {badge(platform_val, PLATFORM_COLORS.get(platform_val, DHL_GRAY))}
                        {badge('Backlog', STATUS_COLORS['Backlog'])}
                        {badge(priority_val, PRIORITY_COLORS.get(priority_val, DHL_GRAY))}
                      </div>
                      <small style="color:{DHL_GRAY}">Logged by: {updated_by}</small>
                    </div>""", unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"GitHub sync failed: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: UPDATE / DELETE TICKET
# ─────────────────────────────────────────────────────────────────────────────
elif page == "Update / Delete Ticket":
    if not user:
        st.warning("You must be logged in to update or delete tickets.")
    elif tickets.empty:
        st.info("No tickets found.")
    else:
        # My Tasks shortcut banner
        if st.session_state.get("my_tasks_mode"):
            st.markdown(f"""
            <div style="background:#FFCC00;border-radius:6px;padding:8px 14px;
                        margin-bottom:12px;font-size:13px;font-weight:700;color:#1A1A1A">
              Showing tickets assigned to you.
            </div>""", unsafe_allow_html=True)
            col_back, _ = st.columns([1,3])
            with col_back:
                if st.button("← Show All Tickets"):
                    st.session_state.my_tasks_mode = False
                    st.session_state.nav_page = "Update / Delete Ticket"
                    st.rerun()

        # Filter by status
        uf1, uf2 = st.columns([1, 2])
        with uf1:
            default_statuses = [s for s in STATUS_ORDER if s != "Done"]
            filter_status = st.multiselect(
                "Filter by Status", STATUS_ORDER,
                default=default_statuses,
                key="update_status_filter"
            )
        with uf2:
            st.markdown("")

        filtered_tickets = tickets[tickets["status"].isin(filter_status)] if filter_status else tickets

        if st.session_state.get("my_tasks_mode") and "assigned_to" in filtered_tickets.columns:
            filtered_tickets = filtered_tickets[filtered_tickets["assigned_to"] == user]

        filtered_tickets = filtered_tickets.sort_values("timestamp", ascending=False)

        if filtered_tickets.empty:
            if st.session_state.get("my_tasks_mode"):
                st.info("No pending tickets assigned to you.")
            else:
                st.info("No tickets match the selected status filter.")
            st.stop()

        options = {f"[{r['status']}] {r['ticket_id']} - {r['title']}": r['ticket_id']
                   for _, r in filtered_tickets.iterrows()}

        # Determine default index — honour jump_to_ticket if set
        default_idx = 0
        jumped_id = st.session_state.get("jump_to_ticket")
        if jumped_id:
            vals = list(options.values())
            if jumped_id in vals:
                default_idx = vals.index(jumped_id)
            st.session_state.jump_to_ticket = None  # clear after use

        sel_label = st.selectbox("Select Ticket", list(options.keys()), index=default_idx)
        sel_id    = options[sel_label]
        t         = tickets[tickets["ticket_id"] == sel_id].iloc[0]

        # Current state card
        pct = int(float(t.get("progress", 0) or 0))
        complexity_badge = badge("Complexity: " + str(t.get("complexity","")), "#555555") if t.get("complexity") else ""
        assigned_badge   = badge("Assigned: " + str(t.get("assigned_to","Unassigned")), "#1A1A1A", DHL_YELLOW) if t.get("assigned_to") else badge("Unassigned", "#9E9E9E")
        st.markdown(
            '<div class="ticket-card">'
            + f'<div class="ticket-id">{t["ticket_id"]} · Requestor: {t.get("requestor","")} · Due: {t.get("due_date","")}</div>'
            + f'<div class="ticket-title">{t["title"]}</div>'
            + '<div class="pills">'
            + badge(t["platform"], PLATFORM_COLORS.get(t["platform"], DHL_GRAY))
            + badge(t["status"],   STATUS_COLORS.get(t["status"],    DHL_GRAY))
            + badge(t["priority"], PRIORITY_COLORS.get(t["priority"], DHL_GRAY))
            + complexity_badge + assigned_badge + '</div>'
            + progress_bar(pct, STATUS_COLORS.get(t["status"], DHL_YELLOW))
            + f'<p style="color:#6B6B6B;font-size:13px;margin-top:6px">{t.get("description","")}</p>'
            + '</div>',
            unsafe_allow_html=True)

        st.markdown("#### Update Fields")
        uc1,uc2,uc3 = st.columns(3)
        cur_status     = t["status"]     if t["status"]     in STATUS_ORDER     else STATUS_ORDER[0]
        cur_priority   = t["priority"]   if t["priority"]   in PRIORITY_ORDER   else PRIORITY_ORDER[0]
        cur_complexity = t.get("complexity","") if t.get("complexity","") in COMPLEXITY_ORDER else COMPLEXITY_ORDER[0]
        with uc1: new_status   = st.selectbox("Status",   STATUS_ORDER,   index=STATUS_ORDER.index(cur_status))
        with uc2: new_priority = st.selectbox("Priority", PRIORITY_ORDER, index=PRIORITY_ORDER.index(cur_priority))
        with uc3: new_progress = st.slider("Progress %", 0, 100, pct, step=5)

        uc4,uc5 = st.columns(2)
        user_list    = list(USERS.keys())
        cur_assigned = t.get("assigned_to","") if t.get("assigned_to","") in user_list else user_list[0]
        with uc4: new_complexity = st.selectbox("Complexity", COMPLEXITY_ORDER,
                                                 index=COMPLEXITY_ORDER.index(cur_complexity))
        with uc5: new_assigned   = st.selectbox("Assign To",  user_list,
                                                 index=user_list.index(cur_assigned) if cur_assigned in user_list else 0)

        nt1, nt2 = st.columns([4, 1])
        with nt1:
            new_notes = st.text_area("Notes / Comment *",
                                     placeholder="Describe what changed or why...")
        with nt2:
            time_taken = st.number_input("Time taken (min)", min_value=5, max_value=480, value=30, step=5)

        if st.button("Save Update"):
            if not new_notes.strip():
                st.error("Please add a note describing the update.")
            else:
                row = {
                    "timestamp":   now8().isoformat(timespec="seconds"),
                    "action":      "UPDATED",
                    "ticket_id":   sel_id,
                    "title":       t["title"],
                    "platform":    t["platform"],
                    "priority":    new_priority,
                    "status":      new_status,
                    "progress":    new_progress,
                    "requestor":   t["requestor"],
                    "due_date":    t["due_date"],
                    "tags":        t["tags"],
                    "description": t["description"],
                    "updated_by":  user,
                    "notes":       new_notes.strip(),
                    "complexity":  new_complexity,
                    "assigned_to": new_assigned,
                }
                with st.spinner("Saving to GitHub..."):
                    try:
                        gh_append(row)
                        act_append({
                            "timestamp":    now8().isoformat(timespec="seconds"),
                            "date":         now8().strftime("%Y-%m-%d"),
                            "username":     user,
                            "category":    "Review",
                            "description":  f"Updated {sel_id} [{t['title']}] → {new_status} {new_progress}% | {new_notes.strip()[:120]}",
                            "duration_min": str(time_taken),
                        })
                        st.success(f"Ticket {sel_id} updated and logged to CSV.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"GitHub sync failed: {e}")

        # Ticket history
        ticket_log = log_df[log_df["ticket_id"] == sel_id].sort_values("timestamp", ascending=False)
        if not ticket_log.empty:
            st.markdown("#### History for this ticket")
            for _, row in ticket_log.iterrows():
                action_color = {"CREATED":"#2E7D32","UPDATED":"#0078D4","DELETED":"#D40511"}
                ac = row["action"]
                st.markdown(f"""
                <div style="background:{DHL_LIGHT};border:1px solid {DHL_BORDER};
                            border-left:4px solid {action_color.get(ac, DHL_GRAY)};
                            border-radius:6px;padding:10px 14px;margin-bottom:8px">
                  <div style="display:flex;gap:12px;align-items:center;margin-bottom:4px">
                    {badge(ac, action_color.get(ac, DHL_GRAY))}
                    <span style="font-size:12px;color:{DHL_GRAY}">{fmt_ts(row['timestamp'])}</span>
                    <span style="font-size:12px;font-weight:700;color:{DHL_DARK}">by {row['updated_by']}</span>
                  </div>
                  <div style="font-size:13px;color:{DHL_DARK}">
                    Status: <b>{row['status']}</b> &nbsp;|&nbsp;
                    Priority: <b>{row['priority']}</b> &nbsp;|&nbsp;
                    Progress: <b>{row['progress']}%</b> &nbsp;|&nbsp;
                    Complexity: <b>{row.get('complexity','—')}</b> &nbsp;|&nbsp;
                    Assigned: <b>{row.get('assigned_to','—')}</b>
                  </div>
                  {f'<div style="font-size:13px;color:{DHL_GRAY};margin-top:4px">{row["notes"]}</div>' if row.get("notes") else ""}
                </div>""", unsafe_allow_html=True)

        # Delete zone
        st.markdown("---")
        with st.expander("Danger Zone"):
            st.warning("This logs a DELETED action. The ticket will be removed from the active view but the full history is preserved in the CSV.")
            del_note = st.text_input("Reason for deletion *")
            if st.button("Delete Ticket", type="secondary"):
                if not del_note.strip():
                    st.error("Please provide a reason for deletion.")
                else:
                    row = {
                        "timestamp":   now8().isoformat(timespec="seconds"),
                        "action":      "DELETED",
                        "ticket_id":   sel_id,
                        "title":       t["title"],
                        "platform":    t["platform"],
                        "priority":    t["priority"],
                        "status":      t["status"],
                        "progress":    t["progress"],
                        "requestor":   t["requestor"],
                        "due_date":    t["due_date"],
                        "tags":        t["tags"],
                        "description": t["description"],
                        "updated_by":  user,
                        "notes":       del_note.strip(),
                    }
                    with st.spinner("Logging deletion..."):
                        try:
                            gh_append(row)
                            act_append({
                                "timestamp":    now8().isoformat(timespec="seconds"),
                                "date":         now8().strftime("%Y-%m-%d"),
                                "username":     user,
                                "category":    "Admin",
                                "description":  f"Deleted {sel_id} [{t['title']}] | Reason: {del_note.strip()[:120]}",
                                "duration_min": "5",
                            })
                            st.success("Ticket deleted from active view. History preserved in CSV.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"GitHub sync failed: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: ACTIVITY LOG
# ─────────────────────────────────────────────────────────────────────────────
elif page == "Activity Log":
    if not user:
        st.warning("You must be logged in to access the Activity Log.")
        st.stop()

    ACT_CATEGORIES = ["Development", "Review", "Meeting", "Support", "Admin", "Other"]
    ACT_COLORS = {
        "Development": "#0078D4",
        "Review":      "#FF8C00",
        "Meeting":     "#6A0DAD",
        "Support":     "#D40511",
        "Admin":       "#2E7D32",
        "Other":       "#6B6B6B",
    }

    act_df = st.session_state.act_df

    with st.expander("➕ Log Today's Activity", expanded=True):
        lc1, lc2, lc3 = st.columns([2, 1, 1])
        with lc1:
            act_desc = st.text_input("What did you work on? *", placeholder="e.g. Fixed Splunk alert threshold for APAC dashboard")
        with lc2:
            act_cat = st.selectbox("Category", ACT_CATEGORIES)
        with lc3:
            act_dur = st.number_input("Duration (min)", min_value=5, max_value=480, value=30, step=5)
        act_date = st.date_input("Date", value=date.today())
        if st.button("Log Activity", use_container_width=False):
            if not act_desc.strip():
                st.error("Please describe what you worked on.")
            else:
                new_act = {
                    "timestamp":    now8().isoformat(timespec="seconds"),
                    "date":         str(act_date),
                    "username":     user,
                    "category":     act_cat,
                    "description":  act_desc.strip(),
                    "duration_min": str(act_dur),
                }
                with st.spinner("Saving..."):
                    try:
                        act_append(new_act)
                        st.success("Activity logged!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Save failed: {e}")

    st.markdown("---")

    import numpy as np

    view_scope = st.radio("View", ["My Activity", "All Team"], horizontal=True)
    heatmap_df = act_df.copy() if not act_df.empty else pd.DataFrame(columns=ACT_COLS)

    if view_scope == "My Activity":
        heatmap_df = heatmap_df[heatmap_df["username"] == user]

    if not heatmap_df.empty:
        heatmap_df["date"] = pd.to_datetime(heatmap_df["date"], errors="coerce")
        heatmap_df = heatmap_df.dropna(subset=["date"])
        heatmap_df["duration_min"] = pd.to_numeric(heatmap_df["duration_min"], errors="coerce").fillna(0)

        st.markdown('<div class="section-header">Contribution Heatmap</div>', unsafe_allow_html=True)

        today      = date.today()
        end_date   = today
        start_date = today - timedelta(days=364)
        date_range = pd.date_range(start=start_date, end=end_date, freq="D")

        daily_counts = heatmap_df.groupby("date")["duration_min"].sum().reset_index()
        daily_counts.columns = ["date", "total_min"]
        daily_counts["date"] = pd.to_datetime(daily_counts["date"])

        date_df = pd.DataFrame({"date": date_range})
        date_df = date_df.merge(daily_counts, on="date", how="left").fillna(0)
        date_df["total_min"] = date_df["total_min"].astype(int)

        pad_start = date_df["date"].iloc[0].weekday()
        padded = pd.concat([
            pd.DataFrame({"date": [None]*pad_start, "total_min": [None]*pad_start}),
            date_df
        ], ignore_index=True)

        num_weeks = int(np.ceil(len(padded) / 7))
        while len(padded) < num_weeks * 7:
            padded = pd.concat([padded, pd.DataFrame({"date":[None],"total_min":[None]})], ignore_index=True)

        grid      = padded["total_min"].values.reshape(num_weeks, 7)
        date_grid = padded["date"].values.reshape(num_weeks, 7)

        max_val = max(date_df["total_min"].max(), 1)

        def intensity_color(v):
            if v is None or (isinstance(v, float) and np.isnan(v)): return "#EEEEEE"
            if v == 0: return "#EEEEEE"
            ratio = v / max_val
            if ratio < 0.25:  return "#C6E48B"
            if ratio < 0.50:  return "#7BC96F"
            if ratio < 0.75:  return "#239A3B"
            return "#196127"

        days_label  = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
        month_labels = []
        last_month   = None
        for w in range(num_weeks):
            for d in range(7):
                cell_date = date_grid[w, d]
                if cell_date is not None and not (isinstance(cell_date, float) and np.isnan(cell_date)):
                    try:
                        dt = pd.Timestamp(cell_date)
                        m  = dt.strftime("%b")
                        if m != last_month:
                            month_labels.append((w, m))
                            last_month = m
                    except:
                        pass

        month_row = [""] * num_weeks
        for (wi, ml) in month_labels:
            month_row[wi] = ml

        month_html = "".join(
            f'<td style="font-size:10px;color:#6B6B6B;text-align:center;padding:0 1px;width:14px">{m}</td>'
            for m in month_row
        )

        rows_html = ""
        for d in range(7):
            cells = ""
            for w in range(num_weeks):
                val       = grid[w, d]
                col       = intensity_color(val)
                tip       = ""
                cell_date = date_grid[w, d]
                if cell_date is not None and not (isinstance(cell_date, float) and np.isnan(cell_date)):
                    try:
                        tip = f"{pd.Timestamp(cell_date).strftime('%d %b %Y')}: {int(val) if val else 0} min"
                    except:
                        tip = ""
                cells += (
                    f'<td title="{tip}" style="width:14px;height:14px;background:{col};'
                    f'border-radius:2px;margin:1px;padding:0"></td>'
                )
            rows_html += (
                f'<tr><td style="font-size:10px;color:#6B6B6B;padding-right:6px;white-space:nowrap">'
                f'{days_label[d]}</td>{cells}</tr>'
            )

        legend_html = "".join(
            f'<span style="display:inline-block;width:12px;height:12px;background:{c};'
            f'border-radius:2px;margin-right:3px;vertical-align:middle"></span>'
            for c in ["#EEEEEE","#C6E48B","#7BC96F","#239A3B","#196127"]
        )

        total_days_logged = int((date_df["total_min"] > 0).sum())
        total_mins        = int(date_df["total_min"].sum())

        st.markdown(f"""
        <div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:20px 24px;margin-bottom:16px;overflow-x:auto">
          <div style="font-size:13px;color:#6B6B6B;margin-bottom:10px">
            <b style="color:#1A1A1A">{total_days_logged}</b> active days &nbsp;·&nbsp;
            <b style="color:#1A1A1A">{total_mins}</b> total minutes logged in the past year
          </div>
          <table style="border-collapse:separate;border-spacing:2px;table-layout:fixed">
            <thead><tr><td></td>{month_html}</tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          <div style="margin-top:10px;font-size:11px;color:#6B6B6B">
            Less {legend_html} More
          </div>
        </div>
        """, unsafe_allow_html=True)

        if view_scope == "All Team":
            st.markdown('<div class="section-header">Team Summary</div>', unsafe_allow_html=True)
            team_sum = (
                heatmap_df.groupby("username")
                .agg(entries=("description","count"), total_min=("duration_min","sum"))
                .reset_index()
                .sort_values("total_min", ascending=False)
            )
            cols = st.columns(min(len(team_sum), 4))
            for i, (_, row) in enumerate(team_sum.iterrows()):
                with cols[i % len(cols)]:
                    hrs  = int(row["total_min"]) // 60
                    mins = int(row["total_min"]) % 60
                    st.markdown(f"""
                    <div class="metric-card">
                      <div class="lbl">{row['username']}</div>
                      <div class="val" style="font-size:1.6rem">{hrs}h {mins}m</div>
                      <div style="font-size:11px;color:#6B6B6B">{int(row['entries'])} entries logged</div>
                    </div>""", unsafe_allow_html=True)

        st.markdown('<div class="section-header">Category Breakdown</div>', unsafe_allow_html=True)
        cat_sum = (
            heatmap_df.groupby("category")["duration_min"].sum()
            .reset_index().sort_values("duration_min", ascending=False)
        )
        if not cat_sum.empty:
            fig_cat = px.bar(
                cat_sum, x="category", y="duration_min",
                color="category", color_discrete_map=ACT_COLORS,
                labels={"duration_min": "Total Minutes", "category": "Category"},
                title="Time by Category",
            )
            fig_cat.update_layout(**CHART)
            st.plotly_chart(fig_cat, use_container_width=True)

        st.markdown('<div class="section-header">Recent Entries</div>', unsafe_allow_html=True)
        show_user   = st.selectbox("Filter by user", ["All"] + sorted(act_df["username"].unique().tolist()), key="act_user_filter") if view_scope == "All Team" else user
        recent_acts = heatmap_df.copy()
        if show_user != "All" and view_scope == "All Team":
            recent_acts = recent_acts[recent_acts["username"] == show_user]
        recent_acts = recent_acts.sort_values("date", ascending=False).head(50)

        for _, row in recent_acts.iterrows():
            cat_col = ACT_COLORS.get(row.get("category","Other"), "#6B6B6B")
            dur     = int(float(row.get("duration_min", 0) or 0))
            hrs_str = f"{dur//60}h {dur%60}m" if dur >= 60 else f"{dur}m"
            st.markdown(f"""
            <div style="background:#fff;border:1px solid #E0E0E0;border-left:4px solid {cat_col};
                        border-radius:6px;padding:10px 14px;margin-bottom:6px;display:flex;
                        justify-content:space-between;align-items:center">
              <div>
                <span style="font-size:11px;color:#6B6B6B;font-weight:700;text-transform:uppercase">
                  {row.get('date','')} · {row.get('username','')}
                </span>
                <div style="font-size:14px;color:#1A1A1A;margin:3px 0">{row.get('description','')}</div>
                {badge(row.get('category','Other'), cat_col)}
              </div>
              <div style="font-size:18px;font-weight:700;color:#1A1A1A;white-space:nowrap;padding-left:16px">{hrs_str}</div>
            </div>""", unsafe_allow_html=True)
    else:
        st.info("No activity logged yet. Use the form above to log your first entry!")
